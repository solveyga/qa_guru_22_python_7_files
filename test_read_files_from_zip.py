import csv
import os
from io import TextIOWrapper
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.pdfgen import canvas

CURRENT_DIR = os.getcwd()
RESOURCES_DIR = os.path.join(CURRENT_DIR, 'resources')
ZIP_FILE = os.path.join(RESOURCES_DIR, 'archive.zip')
TEST_DATA = {
    'csv': [['Header A1', 'Header B1'], ['Cell A2', 'Cell B2']],
    'xlsx': [['Header A1', 'Header B1'], ['Cell A2', 'Cell B2']],
    'pdf': 'Test content\nCell A2  |  Cell B2'
}


@pytest.fixture(scope="module", autouse=True)
def prepare_test_files():
    """Генерирует тестовые файлы *.csv, *.xlsx, *.pdf."""
    if not os.path.exists(RESOURCES_DIR):
        os.mkdir(RESOURCES_DIR)

    csv_path = os.path.join(RESOURCES_DIR, 'csv_file.csv')
    xlsx_path = os.path.join(RESOURCES_DIR, 'xlsx_file.xlsx')
    pdf_path = os.path.join(RESOURCES_DIR, 'pdf_file.pdf')

    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for row in TEST_DATA['csv']:
            writer.writerow(row)

    wb = Workbook()
    ws = wb.active
    for row_data in TEST_DATA['xlsx']:
        ws.append(row_data)
    wb.save(xlsx_path)

    c = canvas.Canvas(pdf_path)
    c.drawString(100, 750, TEST_DATA['pdf'])
    c.save()

    yield

    for path in [csv_path, xlsx_path, pdf_path]:
        if os.path.exists(path):
            os.remove(path)

    if os.path.exists(RESOURCES_DIR) and not os.listdir(RESOURCES_DIR):
        os.rmdir(RESOURCES_DIR)


@pytest.fixture(scope="module", autouse=True)
def create_archive(prepare_test_files):
    """Архивирует тестовые файлы."""

    files = [f for f in os.listdir(RESOURCES_DIR) if f.lower().endswith(('.pdf', '.xlsx', '.csv'))]

    with ZipFile(ZIP_FILE, 'w') as zf:
        for file in files:
            add_file = os.path.join(RESOURCES_DIR, file)
            zf.write(add_file, arcname=file)

    yield

    if os.path.exists(ZIP_FILE):
        os.remove(ZIP_FILE)


def test_csv_in_archive():
    """Проверяет чтение CSV-файла из архива."""
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('csv_file.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file,
                                                      'utf-8-sig')))
            second_row = csvreader[1]

            assert second_row[0] == TEST_DATA['csv'][1][0]
            assert second_row[1] == TEST_DATA['csv'][1][1]


def test_xlsx_in_archive():
    """Проверяет чтение XLSX-файла из архива."""
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('xlsx_file.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell_a2 = sheet.cell(row=2, column=1).value
            cell_b2 = sheet.cell(row=2, column=2).value

            assert cell_a2 == TEST_DATA['xlsx'][1][0]
            assert cell_b2 == TEST_DATA['xlsx'][1][1]


def test_pdf_in_archive():
    """Проверяет чтение PDF-файла из архива."""
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('pdf_file.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

            for part in TEST_DATA['pdf'].split('\n'):
                assert part.strip() in text
